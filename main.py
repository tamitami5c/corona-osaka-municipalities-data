import urllib
import os
import json
import datetime
import re

import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from japanera import Japanera

municipalities=[
    "大阪市","堺市","岸和田市","豊中市","池田市","吹田市","泉大津市","高槻市","貝塚市","守口市",
    "枚方市","茨木市","八尾市","泉佐野市","富田林市","寝屋川市","河内長野市","松原市","大東市",
    "和泉市","箕面市","柏原市","羽曳野市","門真市","摂津市","高石市","藤井寺市","東大阪市",
    "泉南市","四條畷市","交野市","大阪狭山市","阪南市","島本町","豊能町","能勢町","忠岡町",
    "熊取町","田尻町","岬町","太子町","河南町","千早赤阪村"
]
miscellaneous=["大阪府外","調査中","合計"]
# data_keys=municipalities+["大阪府外","調査中","合計"]

def get_file_name(url):
    path=urllib.parse.urlparse(url).path
    return os.path.split(path)[-1]

def download_file(url):
    file_name=get_file_name(url)

    response=requests.get(url)
    if response.status_code!=requests.codes.ok:
        raise Exception("status_code!=200")
    response.encoding=response.apparent_encoding

    with open(file_name,"wb") as f:
        f.write(response.content)
    return file_name

def daterange(start_date,end_date):
    """
    start_date(含む)からend_date(含む)まで
    """
    for n in range(int((end_date-start_date).days)+1):
        yield start_date+datetime.timedelta(n)


def get_latest_info():
    """最新の発表の日付とexcelリンクを取得する
    """
    url="http://www.pref.osaka.lg.jp/iryo/osakakansensho/happyo_kako.html"
    res=requests.get(url)
    res.encoding=res.apparent_encoding
    soup=BeautifulSoup(res.text,"html.parser")
    a_list=soup.select(".detail_free>p>a")
    
    japanera=Japanera()

    #excelファイルのリンクの中で一番新しいものを探す
    latest_date=None
    latest_link=None

    for a in a_list:
        print(a.text)
        if "Excel" in a.text:
            date_text=re.split("[(（]",a.text)[0]
            # date_text=a.text.split("（")[0]
            dt=japanera.strptime(date_text,"%-E%-O年%m月%d日")[0]
            date=dt.date()
            print(date.isoformat())

            if latest_date is None:
                latest_date=date
                latest_link=a.get("href")
            else:
                if date>latest_date:
                    latest_date=date
                    latest_link=a.get("href")

    latest_link=urllib.parse.urljoin(url,latest_link) #絶対パスにする
    print("最新の日付:",latest_date)
    print("リンク:",latest_link)
    return latest_date,latest_link


def main():
    
    date,link=get_latest_info()
    file_name=download_file(link)
    print(link)

    wb=load_workbook(file_name,data_only=True)
    ws=wb["概要1～5"]



    #大阪市のrow番号を求める。位置が固定でなく、日時によって変わってしまうため
    
    target_i=None 
    for row_i in range(1,ws.max_row+1):
        val=ws.cell(row_i,1).value
        if isinstance(val,str):
            if "５　市町村別陽性者発生状況（前日24時まで）" in val:
                target_i=row_i
                break
    start_row=target_i+2 #取得したいテーブルの最初の大阪市の部分は、2セル分だけ下にあるので+2
    print("start_index:",start_row)
    
    #大阪府の市町村ごとの陽性者数を取得する
    d=[]
    d2={}
    #2列分あるので2回する
    for row in range(start_row,start_row+22):
        
        municipality=ws.cell(row,1).value.strip() #市町村名
        count=ws.cell(row,4).value #陽性者数
        total=ws.cell(row,7).value #各市町村ごとの累計陽性者数
        print(municipality,":count=",count,",total=",total)

        if municipality in municipalities:
            d.append({
                "municipality":municipality,
                "count":count,
                "total":total
            })
        elif municipality in miscellaneous:
            d2[municipality]={
                "count":count,
                "total":total
            }
        else:
            raise Exception("not valid municipality:",municipality)


    #テーブルの右側の部分   市町村の部分の長さが↑のものと違う
    for row in range(start_row,start_row+24):
        
        municipality=ws.cell(row,10).value.strip() #市町村名
        count=ws.cell(row,14).value #陽性者数
        total=ws.cell(row,17).value #各市町村ごとの累計陽性者数
        print(municipality,":count=",count,",total=",total)

        if municipality in municipalities:
            d.append({
                "municipality":municipality,
                "count":count,
                "total":total
            })
        elif municipality in miscellaneous:
            d2[municipality]={
                "count":count,
                "total":total
            }
        else:
            raise Exception("not valid municipality:",municipality)
    print(d)
    print(d2)
    print("保存")
    data={
        "date":date.isoformat(),
        # "last_updated":""
        "data":{
            "osaka":d,
            "out_of_osaka":d2["大阪府外"],
            "investigating":d2["調査中"],
            "total":d2["合計"]
        },
    }
    print(data)
    
    with open("osaka_municipalities_data.json","w") as f:
        json.dump(data,f,indent=4,ensure_ascii=False)


if __name__ == "__main__":
    main()